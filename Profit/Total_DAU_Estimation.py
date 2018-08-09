import openpyxl
import os
from itertools import zip_longest
os.chdir('/Users/lezardvaleth/Documents/Python/Profit')

#define how many days the results will show
days = 730
organic = 0
constant = -0.494
registerstage1st = 10000
registerstage2nd = 0
promostage1st = 1
promostage2nd = 1
cpa1st = 2.5
cpa2nd = 0
sever_cost = 0
payment_cost = 0 #FB payment
dev_cost = 0.44


#build a dictionary of retention
temp_r = openpyxl.load_workbook('test.xlsx')
sheet_r = temp_r.get_sheet_by_name('Retention')
retention = {}
for i in range(6):
	retention[sheet_r['A'+str(i+1)].value] = sheet_r['B'+str(i+1)].value

#build a list of arpu
temp_arpu = openpyxl.load_workbook('test.xlsx')
sheet_arpu = temp_arpu.get_sheet_by_name('ARPU')
arpu = []
for i in range(days):
        arpu.append(sheet_arpu['B'+str(i+2)].value)

#calculate single sever DAU and return the DAU
def dau_cal(retention,constant,register):
    dau = []
    dau.append(register)
    for i in range(days):
        dau = dau + [(retention*((i+2)**constant))*register]
    return(dau)

#calculate total DAU depends on the promodays and return the total DAU
def total_dau_cal(retention,constant,registerstage1st,registerstage2nd,promostage1st,promostage2nd):
    templist = []
    total_dau = []
    for i in range(promostage1st):
        dau = [0]*i + dau_cal(retention,constant,registerstage1st+organic)
        templist.append(dau)
    for i in range(promostage1st,promostage2nd):
        dau = [0]*i + dau_cal(retention,constant,registerstage2nd+organic)
        templist.append(dau)
    for i in range(promostage2nd):
        total_dau = [x+y for x,y in zip_longest(total_dau,templist[i],fillvalue=0)]
    return(total_dau)

#wirte the total DAU to an xlsx file
def xlsxwriter_dau(cul):
    r = openpyxl.load_workbook('test.xlsx')
    sheet = r.get_sheet_by_name('DAU')
    for i in range(days):
        sheet[cul+str(1)] = retention[cul]
        sheet[cul+str(i+3)] = total_dau[i]
    r.save('test.xlsx')

#calculate the daily revenue from given arpu
def daily_revenue_cal(given_arpu):
        daily_revenue = []
        for i in range(days):
                daily_revenue.append(total_dau[i] * given_arpu[i])
        return(daily_revenue)
        
#calculate the profit
def profit_cal(sever_cost,payment_cost,dev_cost,cpa1st,cpa2nd,registerstage1st,registerstage2nd,promostage1st,promostage2nd):
        daily_profit = []
        profit = []
        for i in range(promostage1st):
                daily_profit.append(daily_revenue[i]*(1-sever_cost-payment_cost-dev_cost)-cpa1st*registerstage1st)
        for i in range(promostage1st,promostage2nd):
                daily_profit.append(daily_revenue[i]*(1-sever_cost-payment_cost-dev_cost)-cpa2nd*registerstage2nd)
        for i in range(promostage2nd,days):
                daily_profit.append(daily_revenue[i]*(1-sever_cost-payment_cost-dev_cost))
        profit.append(daily_profit[0])
        for i in range(days-1):
                profit.append(profit[i]+daily_profit[i+1])
        return(profit)

#write the daily revenue to an xlsx file
def xlsxwriter_revenue(cul):
        r = openpyxl.load_workbook('test.xlsx')
        sheet = r.get_sheet_by_name('Revenue')
        for i in range(days):
                sheet[cul+str(1)] = retention[cul]
                sheet[cul+str(i+3)] = daily_revenue[i]
        r.save('test.xlsx')

#write the profit to an xlsx file
def xlsxwriter_profit(cul):
        r = openpyxl.load_workbook('test.xlsx')
        sheet = r.get_sheet_by_name('Profit')
        for i in range(days):
                sheet[cul+str(1)] = retention[cul]
                sheet[cul+str(i+3)] = profit[i]
        r.save('test.xlsx')
       
#out put excel files     
for cul,ret in retention.items():
    print(ret)
    print(cul)
    total_dau = total_dau_cal(ret,constant,registerstage1st,registerstage2nd,promostage1st,promostage2nd)
    xlsxwriter_dau(cul)
    daily_revenue = daily_revenue_cal(arpu)
    profit = profit_cal(sever_cost,payment_cost,dev_cost,cpa1st,cpa2nd,registerstage1st,registerstage2nd,promostage1st,promostage2nd)
    xlsxwriter_revenue(cul)
    xlsxwriter_profit(cul)
  

    






    

