#!/usr/bin/env python3
#import and initiate driver
import arrow
import datetime
import time
import openpyxl
from openpyxl import Workbook
wb = Workbook()
import os
from selenium import webdriver
options = webdriver.ChromeOptions()
options.add_argument('headless')
options.add_argument('window-size=800x600')
driver = webdriver.Chrome(executable_path='/Users/lezardvaleth/Documents/Python/chromedriver',chrome_options=options)
os.chdir('/Users/lezardvaleth/Documents/Python/Report')

#use arrow to generate date automaticly
backend_date = arrow.now() - datetime.timedelta(days=2)
year = backend_date.format('YYYY')
month = backend_date.format('MM')
last_month = str(int(month)-1)

#determing date range
time_start = backend_date.format('YYYY-MM')+'-01'
time_end = backend_date.format('YYYY-MM-DD')
time_start_last_month = year+'-'+last_month+'-01'
time_end_last_month = year+'-'+last_month+'-'+backend_date.format('DD')

#time_start = '2017-10-01'
#time_end = '2017-10-31'
#time_start_last_month = '2017-09-01'
#time_end_last_month = '2017-09-30'

#get currency
report = openpyxl.load_workbook('Report.xlsx')
revenue_sheet = report.get_sheet_by_name('Revenue')
currency = []
for i in range(24):
    currency.append(revenue_sheet['D'+str(i+2)].value)


#make a revenue list
revenue = []
revenue_last_month = []

#login in data backend
driver.get('http://acc.r2games.com/main/?r=site/login')
account = driver.find_element_by_name('account')
password = driver.find_element_by_name('password')
login = driver.find_elements_by_id('submit-btn')
account.send_keys('10024')
password.send_keys('yyy666')
account.submit()

#get a list of games
report = openpyxl.load_workbook('Report.xlsx')
revenue_sheet = report.get_sheet_by_name('Revenue')
game_names =[]
for i in range(24):
    game_names.append(revenue_sheet['A'+str(i+2)].value)

#set a list for game inquery
report = openpyxl.load_workbook('Report.xlsx')
revenue_sheet = report.get_sheet_by_name('Revenue')
game_list_en=[]
game_list_fr=[]
game_list_de=[]
for i in range(16):
    game_list_en.append(revenue_sheet['B'+str(i+2)].value)
for i in range(4):
    game_list_fr.append(revenue_sheet['B'+str(i+17)].value)
for i in range(3):
    game_list_de.append(revenue_sheet['B'+str(i+21)].value)

#build a data extractor
def data_extractor(r,m,p,g,server,time_start,time_end,show):

    time.sleep(1)

    driver.get('http://data.r2games.com/platform/?r='+r+'&m='+m+'&p='+p+'&g='+g+'&server='+server+'&time_start='+time_start+'&time_end='+time_end+'&show='+show)
    recharge_temp = driver.find_element_by_xpath('//*[@id="__t_grid_0"]/tbody/tr[1]/td[2]')
    vip = driver.find_element_by_xpath('//*[@id="__t_grid_0"]/tbody/tr[1]/td[9]')

    if vip.text == '':
        recharge= float(recharge_temp.text)
    else:
        recharge=float(recharge_temp.text) + float(vip.text)
    
    return recharge

#get this month revenues
for game in game_list_en:
    revenue.append(data_extractor('user%2Fpayment','M294','r2games',game,'ALL',time_start,time_end,'ByDate'))
    print(game)
for game in game_list_fr:
    revenue.append(data_extractor('user%2Fpayment','M294','fr',game,'ALL',time_start,time_end,'ByDate'))
    print(game)
for game in game_list_de:
    revenue.append(data_extractor('user%2Fpayment','M294','de',game,'ALL',time_start,time_end,'ByDate'))
    print(game)
revenue.append(data_extractor('user%2Fpayment','M294','ru','shadow','ALL',time_start,time_end,'ByDate'))

print('-----------------------------------this month revenue finished-----------------')

#get last month revenues
for game in game_list_en:
    revenue_last_month.append(data_extractor('user%2Fpayment','M294','r2games',game,'ALL',time_start_last_month,time_end_last_month,'ByDate'))
    print(game)
for game in game_list_fr:
    revenue_last_month.append(data_extractor('user%2Fpayment','M294','fr',game,'ALL',time_start_last_month,time_end_last_month,'ByDate'))
    print(game)
for game in game_list_de:
    revenue_last_month.append(data_extractor('user%2Fpayment','M294','de',game,'ALL',time_start_last_month,time_end_last_month,'ByDate'))
    print(game)
revenue_last_month.append(data_extractor('user%2Fpayment','M294','ru','shadow','ALL',time_start_last_month,time_end_last_month,'ByDate'))


#create a new xlsx file
new_report = wb.active
#write inital texts in xlsx file
cell = 0
for name in game_names:
    new_report['A'+str(cell+2)] = name
    cell += 1
new_report['B1'] = time_start
new_report['C1'] = time_start_last_month
new_report['D1'] = 'this month USD revenue'
new_report['E1'] = 'last month USD revenue'
new_report['F1'] = 'changed percentage'
#write data
cell_revenue = 0
for rev in revenue:
    new_report['B'+str(cell_revenue+2)]= rev
    cell_revenue += 1
cell_revenue_last = 0
for rev in revenue_last_month:
    new_report['C'+str(cell_revenue_last+2)]= rev
    cell_revenue_last += 1
cell_revenue = 0
#for rev in revenue:
#    new_report['D'+str(cell_revenue+2)]= rev*currency[cell_revenue]
#    cell_revenue += 1
#cell_revenue_last = 0
#for rev in revenue_last_month:
#    new_report['E'+str(cell_revenue_last+2)]= rev*currency[cell_revenue_last]
#    cell_revenue_last += 1
#for result in range(22):
#    new_report['F'+str(result+2)] = (new_report['D'+str(result+2)].value/new_report['E'+str(result+2)].value) -1

#save
wb.save('report'+' '+time_end_last_month+' and '+time_end+'.xlsx')



#quit browser
driver.quit()
